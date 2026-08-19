import streamlit as st
from pathlib import Path
import pandas as pd
import numpy as np
import io
import openpyxl
import matplotlib.pyplot as plt

# Carpeta raíz donde viven todas las subcarpetas del árbol (Hincado > ... > Excel)
DATA_DIR = Path(__file__).parent / "data"
st.set_page_config(page_title="Ensayos en Suelo a Escala", layout="wide")
st.title("🧪 Ensayos en Suelo a Escala – Acero Corrugado")

if not DATA_DIR.exists():
    st.error(f"No encuentro la carpeta '{DATA_DIR.name}'. Crea una carpeta 'data' junto a app.py "
              "y pon ahí tu árbol de subcarpetas con los Excel.")
    st.stop()


def listar_subcarpetas(path: Path):
    return sorted([p for p in path.iterdir() if p.is_dir()], key=lambda p: p.name)


def listar_archivos(path: Path):
    return sorted(
        [p for p in path.iterdir() if p.suffix.lower() in (".xlsx", ".xls")],
        key=lambda p: p.name,
    )


# --- Funciones para la gráfica combinada suavizada de varios ensayos ---

def _leer_carga_desplazamiento(
    archivo,
    hoja: str = "Hoja1",
    col_carga: int = 5,          # columna E
    col_desplazamiento: int = 7,  # columna G
):
    """
    Lee un archivo .xlsx (ruta o bytes) y devuelve (desplazamiento, carga)
    como arrays de numpy, ordenados por desplazamiento y sin celdas vacías.
    """
    if hasattr(archivo, "read"):
        contenido = archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    else:
        wb = openpyxl.load_workbook(archivo, data_only=True)

    if hoja not in wb.sheetnames:
        raise ValueError(f"La hoja '{hoja}' no existe en este archivo.")

    ws = wb[hoja]

    desplazamiento, carga = [], []
    for fila in range(1, ws.max_row + 1):
        d = ws.cell(row=fila, column=col_desplazamiento).value
        c = ws.cell(row=fila, column=col_carga).value
        if d is None or c is None:
            continue
        desplazamiento.append(d)
        carga.append(c)

    desplazamiento = np.array(desplazamiento, dtype=float)
    carga = np.array(carga, dtype=float)

    orden = np.argsort(desplazamiento)
    return desplazamiento[orden], carga[orden]


def _promedio_movil(y: np.ndarray, ventana: int = 5) -> np.ndarray:
    """Suaviza una serie con un promedio móvil centrado de tamaño `ventana`."""
    if ventana <= 1:
        return y
    kernel = np.ones(ventana) / ventana
    y_pad = np.pad(y, (ventana // 2, ventana // 2), mode="edge")
    return np.convolve(y_pad, kernel, mode="valid")[: len(y)]


def graficar_ensayos_combinados(
    archivos,
    nombres=None,
    ventana_suavizado: int = 5,
    hoja: str = "Hoja1",
    col_carga: int = 5,
    col_desplazamiento: int = 7,
    titulo: str = "Carga vs. Desplazamiento",
    xlabel: str = "Desplazamiento (mm)",
    ylabel: str = "Carga (kN)",
):
    """
    Genera una figura de matplotlib con las curvas suavizadas de N ensayos
    superpuestas. Retorna (fig, errores) donde errores es una lista de
    (nombre, mensaje) para los archivos que no se pudieron leer.
    """
    if nombres is None:
        nombres = [f"Ensayo {i+1}" for i in range(len(archivos))]

    colores = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4", "#4a3aa7"]

    fig, ax = plt.subplots(figsize=(9, 5.5))
    errores = []

    for i, (archivo, nombre) in enumerate(zip(archivos, nombres)):
        try:
            x, y = _leer_carga_desplazamiento(
                archivo, hoja=hoja, col_carga=col_carga, col_desplazamiento=col_desplazamiento
            )
            if len(x) == 0:
                errores.append((nombre, "No se encontraron datos válidos."))
                continue
            y_suave = _promedio_movil(y, ventana=ventana_suavizado)
            ax.plot(x, y_suave, label=nombre, color=colores[i % len(colores)], linewidth=2)
        except Exception as e:
            errores.append((nombre, str(e)))

    ax.set_title(titulo)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.grid(True, color="#e1e0d9", linewidth=0.8)
    ax.legend(frameon=False)
    fig.tight_layout()

    return fig, errores


# --- Navegación por árbol de carpetas ---

current = DATA_DIR
breadcrumb = []

nivel = 0
while True:
    subfolders = listar_subcarpetas(current)
    if not subfolders:
        break
    etiqueta = "Categoría" if nivel == 0 else current.name.replace("_", " ")
    opciones = [f.name.replace("_", " ") for f in subfolders]
    seleccion = st.selectbox(f"Nivel {nivel + 1} — {etiqueta}", opciones, key=f"nivel_{nivel}")
    current = subfolders[opciones.index(seleccion)]
    breadcrumb.append(seleccion)
    nivel += 1

if breadcrumb:
    st.caption(" ➜ ".join(breadcrumb))

archivos = listar_archivos(current)

if not archivos:
    st.info("No hay archivos Excel en esta rama del árbol todavía.")
else:
    nombres = [f.name for f in archivos]

    # --- Vista de un solo archivo (comportamiento original) ---
    archivo_sel = st.selectbox("Archivo", nombres)
    ruta = archivos[nombres.index(archivo_sel)]

    try:
        hojas = pd.read_excel(ruta, sheet_name=None)
        hoja_sel = st.selectbox("Hoja", list(hojas.keys())) if len(hojas) > 1 else list(hojas.keys())[0]
        df = hojas[hoja_sel]
        st.dataframe(df, use_container_width=True)

        # --- Sección de gráficas (un solo archivo) ---
        columnas_numericas = df.select_dtypes(include="number").columns.tolist()
        todas_columnas = df.columns.tolist()

        if len(todas_columnas) >= 2:
            st.markdown("### 📈 Graficar")
            col1, col2, col3 = st.columns(3)
            with col1:
                eje_x = st.selectbox("Eje X", todas_columnas, key="eje_x")
            with col2:
                opciones_y = [c for c in columnas_numericas if c != eje_x] or columnas_numericas
                eje_y = st.selectbox("Eje Y", opciones_y, key="eje_y")
            with col3:
                tipo = st.selectbox("Tipo de gráfica", ["Línea", "Dispersión", "Barras"], key="tipo_grafica")

            datos_grafica = df[[eje_x, eje_y]].dropna().sort_values(by=eje_x)

            if tipo == "Línea":
                st.line_chart(datos_grafica, x=eje_x, y=eje_y)
            elif tipo == "Dispersión":
                st.scatter_chart(datos_grafica, x=eje_x, y=eje_y)
            else:
                st.bar_chart(datos_grafica, x=eje_x, y=eje_y)
        else:
            st.caption("Esta hoja no tiene suficientes columnas para graficar.")

    except Exception as e:
        st.warning(f"No pude previsualizar el archivo ({e}), pero sí puedes descargarlo.")

    with open(ruta, "rb") as f:
        st.download_button("⬇️ Descargar Excel", f, file_name=archivo_sel)

    # --- Comparar varios ensayos en una sola gráfica suavizada ---
    st.markdown("---")
    st.markdown("### 🔗 Comparar varios ensayos (curva suavizada)")
    st.caption(
        "Selecciona 2 o más archivos de esta misma carpeta para superponerlos en una sola "
        "gráfica de Carga vs. Desplazamiento, igual a la hoja 'Hoja1' de cada Excel."
    )

    archivos_comparar = st.multiselect(
        "Archivos a comparar",
        nombres,
        default=nombres[: min(3, len(nombres))],
        key="archivos_comparar",
    )

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        hoja_datos = st.text_input("Nombre de la hoja con los datos", value="Hoja1", key="hoja_datos")
    with col_b:
        col_carga_letra = st.text_input("Columna de carga (kN)", value="E", key="col_carga")
    with col_c:
        col_desp_letra = st.text_input("Columna de desplazamiento (mm)", value="G", key="col_desp")

    ventana = st.slider("Suavizado (ventana del promedio móvil)", 1, 15, 5, key="ventana_suavizado")

    if st.button("Generar gráfica combinada", key="btn_combinar"):
        if len(archivos_comparar) < 2:
            st.warning("Selecciona al menos 2 archivos.")
        else:
            rutas_sel = [archivos[nombres.index(n)] for n in archivos_comparar]
            col_carga_idx = openpyxl.utils.column_index_from_string(col_carga_letra.strip().upper())
            col_desp_idx = openpyxl.utils.column_index_from_string(col_desp_letra.strip().upper())

            fig, errores = graficar_ensayos_combinados(
                rutas_sel,
                nombres=archivos_comparar,
                ventana_suavizado=ventana,
                hoja=hoja_datos.strip(),
                col_carga=col_carga_idx,
                col_desplazamiento=col_desp_idx,
            )

            st.pyplot(fig)

            if errores:
                for nombre, msg in errores:
                    st.warning(f"{nombre}: {msg}")
